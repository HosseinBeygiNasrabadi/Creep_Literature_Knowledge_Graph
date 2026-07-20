#!/usr/bin/env python3
"""
xlsm2json.py -- convert the creep spreadsheet (single-header template) into
ONE JSON array file: one object per data row.

Usage:
    python3 xlsm2json.py creep_spreadsheet.xlsm [output.json] [--sheet NAME]

Spreadsheet layout:
    row 1:  header keys (human-readable; normalized to snake_case keys here)
    row 2+: one dataset per row; column 'ID' is mandatory and unique

Cell conventions:
    '625 MPa' / '650°C' / '170 µm'  -> {"value": 625.0, "unit": "MPa", "unit_iri": <QUDT>}
    '4'                             -> {"value": 4.0}
    '159.74 ±19.18 h', '< 0.02',
    descriptive text                -> {"text": "..."}
    empty cell                      -> key absent (no triples generated)

This script contains NO ontology semantics: classes live in the YARRRML file,
keyed by the snake_case keys produced here. A new spreadsheet column is
auto-converted (header -> snake_case key); it only needs a matching mapping
block in the YARRRML file.
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

QUDT = "http://qudt.org/vocab/unit/"

# ---- unit token -> QUDT IRI (case-insensitive; purely syntactic, no domain logic) ----
UNIT_MAP = {
    "mpa": QUDT + "MegaPA",
    "gpa": QUDT + "GigaPA",
    "kpa": QUDT + "KiloPA",
    "pa": QUDT + "PA",
    "ksi": QUDT + "KiloPSI",
    "°c": QUDT + "DEG_C",
    "c": QUDT + "DEG_C",
    "k": QUDT + "K",
    "h": QUDT + "HR",
    "hr": QUDT + "HR",
    "min": QUDT + "MIN",
    "s": QUDT + "SEC",
    "1/s": QUDT + "PER-SEC",
    "1/h": QUDT + "PER-HR",
    "%": QUDT + "PERCENT",
    "µm": QUDT + "MicroM",
    "μm": QUDT + "MicroM",   # micro sign vs greek mu
    "um": QUDT + "MicroM",
    "nm": QUDT + "NanoM",
    "mm": QUDT + "MilliM",
    "m": QUDT + "M",
    "kj/mol": QUDT + "KiloJ-PER-MOL",
    "j/mol": QUDT + "J-PER-MOL",
    "ppm": QUDT + "PPM",
    "wt.%": QUDT + "PERCENT",
}

# headers that are handled specially (identification, not parameters)
ID_KEY = "id"
DOI_KEY = "paper_doi"

# normalized-header -> canonical key (covers template wording, typos, synonyms);
# any header not listed here is auto-converted to snake_case
HEADER_ALIASES = {
    "id": "id",
    "paper_doi": "paper_doi",
    "material_name": "material_name",
    "material_chemical_composition": "chemical_composition",
    "chemical_composition": "chemical_composition",
    "sample_id": "sample_id",
    "processing_method": "processing_method",
    "solutionizing": "solutionizing",
    "aging": "aging",
    "grain_size": "grain_size",
    "delta_precepitates_fraction": "delta_precipitate_fraction",
    "delta_precipitate_fraction": "delta_precipitate_fraction",
    "delta_precepitates_size": "delta_precipitate_size",
    "delta_precipitate_size": "delta_precipitate_size",
    "gama_precepitates_fraction": "gamma_precipitate_fraction",
    "gamma_precipitate_fraction": "gamma_precipitate_fraction",
    "gama_precepitates_size": "gamma_precipitate_size",
    "gamma_precipitate_size": "gamma_precipitate_size",
    "gama_prime_precepitates_fraction": "gamma_prime_precipitate_fraction",
    "gamma_prime_precipitate_fraction": "gamma_prime_precipitate_fraction",
    "gama_prime_precepitates_size": "gamma_prime_precipitate_size",
    "gamma_prime_precipitate_size": "gamma_prime_precipitate_size",
    "test_standard": "test_standard",
    "temperature": "temperature",
    "initial_stress": "initial_stress",
    "creep_rupture_time": "creep_rupture_time",
    "percentage_permanent_elongation": "percentage_permanent_elongation",
    "percentage_elongation_after_creep_fracture": "percentage_elongation_after_creep_fracture",
    "percentage_reduction_of_area_after_creep_fracture": "percentage_reduction_of_area",
    "percentage_reduction_of_area": "percentage_reduction_of_area",
    "percentage_total_extension": "percentage_total_extension",
    "percentage_initial_total_extension": "percentage_initial_total_extension",
    "percentage_elastic_extension": "percentage_elastic_extension",
    "percentage_initial_plastic_extension": "percentage_initial_plastic_extension",
    "percentage_plastic_extension": "percentage_plastic_extension",
    "percentage_creep_extension": "percentage_creep_extension",
    "steady_state_creep_rate": "steady_state_creep_rate",
    "stress_exponent_n": "stress_exponent",
    "stress_exponent": "stress_exponent",
    "activation_energy_qc": "activation_energy",
    "activation_energy": "activation_energy",
}

SKIP_SHEETS = {"readme", "archived"}
EMPTY_VALUES = {"", "not applicable", "n/a", "na", "-", "--"}

# strict quantity pattern: '<number>' or '<number> <unit>' (space optional before
# non-alphanumeric unit starts like ° or %). Anything else -> text literal.
QUANTITY_RE = re.compile(
    r"^(-?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?)\s*([^\s].*)?$"
)


def norm_header(h):
    """'Stress Exponent (n)' -> 'stress_exponent_n'; ' Grain size ' -> 'grain_size'."""
    s = str(h).strip().lower()
    s = re.sub(r"[()]", " ", s)
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")


def parse_cell(raw):
    """Return a dict for the JSON: {value[,unit,unit_iri]} or {text}."""
    if isinstance(raw, (int, float)):
        return {"value": float(raw)}
    s = str(raw).strip()
    # anything qualified, uncertain, or exotic stays a text literal (as agreed)
    if any(tok in s for tok in ("±", "<", ">", "~", "×", "…")):
        return {"text": s}
    m = QUANTITY_RE.match(s)
    if not m:
        return {"text": s}
    value = float(m.group(1))
    unit = (m.group(2) or "").strip()
    if not unit:
        return {"value": value}
    # unit token must not itself contain digits/spaces beyond a simple token like '1/s'
    if re.search(r"\d", unit) and unit.lower() not in UNIT_MAP:
        return {"text": s}
    entry = {"value": value, "unit": unit}
    iri = UNIT_MAP.get(unit.lower())
    if iri:
        entry["unit_iri"] = iri
    else:
        print(f"  warning: unit '{unit}' not in QUDT map (kept as plain label)")
    return entry


def pick_sheet(wb, requested=None):
    if requested:
        return wb[requested]
    for name in wb.sheetnames:
        if name.strip().lower() not in SKIP_SHEETS:
            return wb[name]
    raise SystemExit("no usable data sheet found")


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    xlsm = args[0] if args else "creep_spreadsheet.xlsm"
    out_path = Path(args[1]) if len(args) > 1 else Path("input.json")
    sheet = None
    for i, a in enumerate(sys.argv):
        if a == "--sheet" and i + 1 < len(sys.argv):
            sheet = sys.argv[i + 1]

    wb = openpyxl.load_workbook(xlsm, data_only=True)
    ws = pick_sheet(wb, sheet)
    print(f"reading sheet '{ws.title}' ({ws.max_row} rows)")

    # header row -> canonical keys
    keys = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        if h is None:
            continue
        nh = norm_header(h)
        keys[c] = HEADER_ALIASES.get(nh, nh)

    records, seen_ids = [], set()
    for r in range(2, ws.max_row + 1):
        row = {}
        for c, key in keys.items():
            v = ws.cell(r, c).value
            if v is None:
                continue
            if isinstance(v, str) and v.strip().lower() in EMPTY_VALUES:
                continue
            row[key] = v
        if ID_KEY not in row:
            continue

        rid = re.sub(r"[^A-Za-z0-9._-]", "_", str(row.pop(ID_KEY)).strip())
        if rid in seen_ids:
            print(f"  warning: duplicate id '{rid}' (row {r}) -- skipped")
            continue
        seen_ids.add(rid)

        record = {"id": rid}
        doi = row.pop(DOI_KEY, None)
        if doi:
            doi = str(doi).strip()
            record["doi"] = doi if doi.startswith("http") else f"https://doi.org/{doi}"

        for key, raw in row.items():
            entry = parse_cell(raw)
            # row id inside every parameter object: YARRRML subject templates
            # use $(param.id), so absent parameters generate NO triples
            entry["id"] = rid
            record[key] = entry
        records.append(record)

    out_path.write_text(json.dumps(records, indent=2, ensure_ascii=False))
    print(f"{len(records)} record(s) -> {out_path}")


if __name__ == "__main__":
    main()
